from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


MONEY_FORMAT = 'R$ #,##0.00'
PERCENT_FORMAT = '0.00%'
INTEGER_FORMAT = '0'


def _display_value(value: Any, value_type: str) -> str:
    if value is None or value == "":
        return "-"
    if value_type == "money":
        amount = f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"R$ {amount}"
    if value_type == "percent":
        return f"{float(value):.2f}%".replace(".", ",")
    if value_type == "boolean":
        return "Sim" if bool(value) else "Não"
    return str(value)


def _excel_value(value: Any, value_type: str) -> Any:
    if value is None:
        return ""
    if value_type in {"money", "number", "percent"}:
        return float(value)
    if value_type == "integer":
        return int(value)
    if value_type == "boolean":
        return "Sim" if bool(value) else "Não"
    return value


def build_report_xlsx(report: dict) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatório"
    columns = report.get("columns") or []
    rows = report.get("rows") or []
    metadata = report.get("metadata") or {}
    summary = report.get("summary") or []

    sheet["A1"] = str(report.get("title") or "Relatório")
    sheet["A1"].font = Font(size=16, bold=True, color="172235")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(columns)))

    current_row = 3
    for label, value in metadata.get("display", []):
        sheet.cell(current_row, 1, str(label)).font = Font(bold=True)
        sheet.cell(current_row, 2, str(value))
        current_row += 1

    if summary:
        current_row += 1
        for item in summary:
            sheet.cell(current_row, 1, str(item.get("label") or "")).font = Font(bold=True)
            cell = sheet.cell(
                current_row,
                2,
                _excel_value(item.get("value"), str(item.get("type") or "text")),
            )
            value_type = str(item.get("type") or "text")
            if value_type == "money":
                cell.number_format = MONEY_FORMAT
            elif value_type == "percent":
                cell.number_format = '0.00"%"'
            elif value_type == "integer":
                cell.number_format = INTEGER_FORMAT
            current_row += 1

    current_row += 1
    header_row = current_row
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(header_row, index, str(column.get("label") or column.get("key") or ""))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="EF3B73")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        current_row += 1
        for index, column in enumerate(columns, start=1):
            value_type = str(column.get("type") or "text")
            cell = sheet.cell(
                current_row,
                index,
                _excel_value(row.get(column.get("key")), value_type),
            )
            if value_type == "money":
                cell.number_format = MONEY_FORMAT
            elif value_type == "percent":
                cell.number_format = '0.00"%"'
            elif value_type == "integer":
                cell.number_format = INTEGER_FORMAT
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(max(1, len(columns)))}{max(header_row, current_row)}"
    )
    for index, column in enumerate(columns, start=1):
        label_length = len(str(column.get("label") or ""))
        sample_lengths = [
            len(_display_value(row.get(column.get("key")), str(column.get("type") or "text")))
            for row in rows[:200]
        ]
        sheet.column_dimensions[get_column_letter(index)].width = min(
            45,
            max(12, label_length + 2, max(sample_lengths, default=0) + 2),
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_page(canvas, document, report: dict) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    page_number = canvas.getPageNumber()
    canvas.drawRightString(
        landscape(A4)[0] - 12 * mm,
        8 * mm,
        f"Mova Sports | {report.get('title', 'Relatório')} | Página {page_number}",
    )
    canvas.restoreState()


def build_report_pdf(report: dict) -> bytes:
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=14 * mm,
        title=str(report.get("title") or "Relatório"),
        author="Mova Sports",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=19,
        textColor=colors.HexColor("#172235"),
        alignment=TA_LEFT,
        spaceAfter=5 * mm,
    )
    meta_style = ParagraphStyle(
        "ReportMeta",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#5D6B82"),
    )
    cell_style = ParagraphStyle(
        "ReportCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=6.4,
        leading=8,
        alignment=TA_LEFT,
    )
    header_style = ParagraphStyle(
        "ReportHeader",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=colors.white,
        alignment=TA_CENTER,
    )

    story = [
        Paragraph("MOVA SPORTS", title_style),
        Paragraph(str(report.get("title") or "Relatório"), styles["Heading2"]),
    ]
    for label, value in (report.get("metadata") or {}).get("display", []):
        story.append(Paragraph(f"<b>{label}:</b> {value}", meta_style))
    story.append(Spacer(1, 4 * mm))

    summary = report.get("summary") or []
    if summary:
        summary_data = [[
            Paragraph(str(item.get("label") or ""), header_style)
            for item in summary
        ], [
            Paragraph(
                _display_value(item.get("value"), str(item.get("type") or "text")),
                cell_style,
            )
            for item in summary
        ]]
        summary_table = Table(summary_data, repeatRows=1)
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EF3B73")),
            ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 1), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.extend((summary_table, Spacer(1, 5 * mm)))

    columns = report.get("columns") or []
    rows = report.get("rows") or []
    table_data = [[
        Paragraph(str(column.get("label") or column.get("key") or ""), header_style)
        for column in columns
    ]]
    for row in rows:
        table_data.append([
            Paragraph(
                _display_value(row.get(column.get("key")), str(column.get("type") or "text")),
                cell_style,
            )
            for column in columns
        ])
    if len(table_data) == 1:
        table_data.append([
            Paragraph("Nenhum registro encontrado para os filtros informados.", cell_style),
            *[Paragraph("", cell_style) for _ in columns[1:]],
        ])

    available_width = landscape(A4)[0] - 20 * mm
    widths = [available_width / max(1, len(columns))] * max(1, len(columns))
    report_table = Table(table_data, colWidths=widths, repeatRows=1)
    report_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#172235")),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), (colors.white, colors.HexColor("#F8FAFC"))),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(report_table)

    document.build(
        story,
        onFirstPage=lambda canvas, doc: _pdf_page(canvas, doc, report),
        onLaterPages=lambda canvas, doc: _pdf_page(canvas, doc, report),
    )
    return output.getvalue()
